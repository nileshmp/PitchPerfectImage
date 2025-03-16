import csv
import openpyxl
from logger_config import logger

class CSV:
    def parse(self, file_path):
        logger.debug('CSV file path is %s', file_path)
        with open(file_path, newline='') as csvfile:
            spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
            for row in spamreader:
                print(', '.join(row))

class XLSX:
    def parse(self, file_path):
        logger.debug('XLSX file path is %s', file_path)
        dataframe = openpyxl.load_workbook(file_path)
        # Define variable to read sheet
        dataframe1 = dataframe.active
        # Iterate the loop to read the cell values
        videos = []
        for row in range(1, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                videos.append(col[row].value)
        return videos